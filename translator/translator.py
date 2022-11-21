from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from googletrans import Translator
from deep_translator import GoogleTranslator

translator = Translator()

workbook = load_workbook(filename="DATA2020_midbc_rubrics_ey.xlsx")

sheet = workbook[workbook.sheetnames[0]]

src='en'
dest = 'es'

sheet_copy = workbook.copy_worksheet(sheet)

from copy import copy

for row in sheet.rows:
    for cell in row:
        try:
            translated_text = GoogleTranslator(source='auto', target='es').translate(cell.value)
            (r, c) = coordinate_to_tuple(cell.coordinate)
            new_cell = sheet_copy.cell(row=r, column=c, value= translated_text)
            print(f"Completed cell: {cell.coordinate}. Text: {translated_text}")
        except:
            print(cell.value)
            pass

workbook.save('prueba.xlsx')